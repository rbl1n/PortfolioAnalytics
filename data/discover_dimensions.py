"""Extract all unique attribute values from monthly sheets to discover potential filter dimensions."""
import openpyxl
import os
import json

FILEPATH = os.path.join(os.path.dirname(__file__), "raw", "理財彙總2025_珊.xlsx")
OUTPUT = os.path.join(os.path.dirname(__file__), "dimension_discovery.txt")

def discover_dimensions():
    wb = openpyxl.load_workbook(FILEPATH, data_only=True)
    target_sheets = [s for s in wb.sheetnames if s.startswith("各行理財存款清單")]
    
    # Use the most complete sheet (1231) for dimension discovery
    ws = wb["各行理財存款清單-1231"]
    
    # Find header row
    header_row = 5  # Based on analysis
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val:
            headers[col] = str(val).strip()
    
    # Collect all data with bank grouping
    current_bank = None
    records = []
    
    for row_idx in range(4, ws.max_row + 1):
        # Check if this is a bank header row
        col1 = ws.cell(row=row_idx, column=1).value
        col2 = ws.cell(row=row_idx, column=2).value
        
        # Bank header detection: col1 has text but col2 is empty, and it's not a cert number
        if col1 and not col2 and not str(col1).startswith("00") and str(col1) != "憑證號碼":
            # Check if this is a bank name (not a summary row)
            if "合計" not in str(col1) and "粗估" not in str(col1):
                current_bank = str(col1).strip()
                continue
        
        # Skip header rows
        if col1 and str(col1) == "憑證號碼":
            continue
            
        # Data row - must have cert number starting with 00
        if col1 and str(col1).startswith("00"):
            record = {"bank": current_bank}
            for col_idx, header_name in headers.items():
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    record[header_name] = val
            records.append(record)
    
    wb.close()
    
    # Analyze unique values for each dimension
    with open(OUTPUT, "w", encoding="utf-8") as f:
        f.write(f"=== Dimension Discovery (from 1231 sheet) ===\n")
        f.write(f"Total records: {len(records)}\n\n")
        
        # Unique banks
        banks = set(r.get("bank", "N/A") for r in records)
        f.write(f"--- Banks ({len(banks)}) ---\n")
        for b in sorted(banks, key=lambda x: str(x)):
            count = sum(1 for r in records if r.get("bank") == b)
            f.write(f"  {b}: {count} funds\n")
        
        # Unique types (型態)
        f.write(f"\n--- Types (型態) ---\n")
        types = {}
        for r in records:
            t = str(r.get("型態", "N/A"))
            types[t] = types.get(t, 0) + 1
        for t, c in sorted(types.items()):
            f.write(f"  {t}: {c}\n")
        
        # Unique currencies (幣別)
        f.write(f"\n--- Currencies (幣別) ---\n")
        currencies = {}
        for r in records:
            curr = str(r.get("幣別", "N/A"))
            currencies[curr] = currencies.get(curr, 0) + 1
        for c, n in sorted(currencies.items()):
            f.write(f"  {c}: {n}\n")
        
        # Dividend types (配息)
        f.write(f"\n--- Dividend Types (配息) ---\n")
        divs = {}
        for r in records:
            d = str(r.get("配息", "N/A"))
            divs[d] = divs.get(d, 0) + 1
        for d, n in sorted(divs.items()):
            f.write(f"  {d}: {n}\n")
        
        # Fee types (費用說明)
        f.write(f"\n--- Fee Types (費用說明) ---\n")
        fees = {}
        for r in records:
            fee = str(r.get("費用說明", "N/A"))
            fees[fee] = fees.get(fee, 0) + 1
        for fee, n in sorted(fees.items()):
            f.write(f"  {fee}: {n}\n")
        
        # Investment start year distribution
        f.write(f"\n--- Investment Start Year ---\n")
        years = {}
        for r in records:
            d = r.get("投資始日")
            if d:
                try:
                    y = str(d)[:4]
                    years[y] = years.get(y, 0) + 1
                except:
                    pass
        for y, n in sorted(years.items()):
            f.write(f"  {y}: {n}\n")
        
        # Investment duration ranges (calculate from 投資始日 to 2025-12-31)
        f.write(f"\n--- Investment Duration (years, approx) ---\n")
        from datetime import datetime
        end_date = datetime(2025, 12, 31)
        durations = {"<1年": 0, "1-3年": 0, "3-5年": 0, "5-10年": 0, ">10年": 0}
        for r in records:
            d = r.get("投資始日")
            if d:
                try:
                    if isinstance(d, datetime):
                        delta = (end_date - d).days / 365.25
                    else:
                        continue
                    if delta < 1: durations["<1年"] += 1
                    elif delta < 3: durations["1-3年"] += 1
                    elif delta < 5: durations["3-5年"] += 1
                    elif delta < 10: durations["5-10年"] += 1
                    else: durations[">10年"] += 1
                except:
                    pass
        for d, n in sorted(durations.items()):
            f.write(f"  {d}: {n}\n")
        
        # Return rate ranges (含息報酬%)
        f.write(f"\n--- Return Rate Distribution (含息報酬%) ---\n")
        returns = {"虧損(<0%)": 0, "微利(0-5%)": 0, "中等(5-20%)": 0, "良好(20-50%)": 0, "優秀(>50%)": 0}
        for r in records:
            ret = r.get("含息報酬%")
            if ret is not None:
                try:
                    ret = float(ret)
                    if ret < 0: returns["虧損(<0%)"] += 1
                    elif ret < 5: returns["微利(0-5%)"] += 1
                    elif ret < 20: returns["中等(5-20%)"] += 1
                    elif ret < 50: returns["良好(20-50%)"] += 1
                    else: returns["優秀(>50%)"] += 1
                except:
                    pass
        for r, n in sorted(returns.items()):
            f.write(f"  {r}: {n}\n")
        
        # Fund name keywords for categorization
        f.write(f"\n--- Fund Categories (by name keywords) ---\n")
        categories = {
            "債券型": ["債", "Bond"],
            "股票型": ["股", "Stock", "Equity"],
            "多重資產": ["多重資產", "Multi"],
            "特別股": ["特別股"],
            "黃金": ["黃金", "Gold"],
            "醫療": ["醫療", "Health"],
            "保險": ["保險", "壽"],
            "永續/ESG": ["永續", "ESG"],
            "政府債": ["政府"],
            "高收益債": ["非投資等級", "High Yield"],
        }
        for cat, keywords in categories.items():
            count = 0
            examples = []
            for r in records:
                name = str(r.get("基金名稱", ""))
                if any(kw in name for kw in keywords):
                    count += 1
                    if len(examples) < 2:
                        examples.append(name[:30])
            f.write(f"  {cat}: {count} (e.g. {', '.join(examples)})\n")
        
        # Monthly trend direction (上月含息損益 vs 含息損益)
        f.write(f"\n--- Monthly Trend Direction ---\n")
        trend = {"上升": 0, "持平": 0, "下降": 0, "無資料": 0}
        for r in records:
            current = r.get("含息損益")
            prev = r.get("上月含息損益")
            diff = r.get("含息利益兩月比較")
            if diff is not None:
                try:
                    diff = float(diff)
                    if diff > 0: trend["上升"] += 1
                    elif diff == 0: trend["持平"] += 1
                    else: trend["下降"] += 1
                except:
                    trend["無資料"] += 1
            else:
                trend["無資料"] += 1
        for t, n in sorted(trend.items()):
            f.write(f"  {t}: {n}\n")
        
        # All unique fund names
        f.write(f"\n--- All Fund Names ({len(records)} total) ---\n")
        for r in sorted(records, key=lambda x: str(x.get("bank", "")) + str(x.get("基金名稱", ""))):
            f.write(f"  [{r.get('bank')}] {r.get('基金名稱', 'N/A')[:40]} | {r.get('幣別','?')} | {r.get('型態','?')} | 配息:{r.get('配息','?')}\n")
    
    print(f"Dimension discovery written to {OUTPUT}")

if __name__ == "__main__":
    discover_dimensions()
