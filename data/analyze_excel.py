"""Analyze Excel files in data/raw/ and output structure to a text report."""
import openpyxl
import os
import json

RAW_DIR = os.path.join(os.path.dirname(__file__), "raw")
OUTPUT_FILE = os.path.join(os.path.dirname(__file__), "analysis_report.txt")

def analyze_workbook(filepath):
    """Analyze a single Excel workbook."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    result = {
        "filename": os.path.basename(filepath),
        "sheets": []
    }
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_info = {
            "name": sheet_name,
            "rows": ws.max_row,
            "cols": ws.max_column,
            "headers": [],
            "sample_rows": [],
            "merged_cells": [str(m) for m in ws.merged_cells.ranges] if ws.merged_cells else []
        }

        # Get headers (first 3 rows to catch multi-level headers)
        for row_idx in range(1, min(4, ws.max_row + 1)):
            row_data = []
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                val = cell.value
                if val is not None:
                    row_data.append(f"[{col_idx}] {val}")
                else:
                    row_data.append(f"[{col_idx}] (empty)")
            sheet_info["headers"].append(row_data)

        # Get sample data rows (rows 4-8 or whatever is available)
        start_row = 4
        for row_idx in range(start_row, min(start_row + 5, ws.max_row + 1)):
            row_data = []
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                val = cell.value
                if val is not None:
                    row_data.append(f"[{col_idx}] {val}")
                else:
                    row_data.append(f"[{col_idx}] (empty)")
            sheet_info["sample_rows"].append(row_data)

        # Get last few rows to see footer/summary patterns
        sheet_info["last_rows"] = []
        if ws.max_row > 8:
            for row_idx in range(max(ws.max_row - 2, 1), ws.max_row + 1):
                row_data = []
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    val = cell.value
                    if val is not None:
                        row_data.append(f"[{col_idx}] {val}")
                    else:
                        row_data.append(f"[{col_idx}] (empty)")
                sheet_info["last_rows"].append(row_data)

        result["sheets"].append(sheet_info)
    
    wb.close()
    return result

def write_report(results):
    """Write analysis report to text file."""
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        for result in results:
            f.write(f"{'='*80}\n")
            f.write(f"FILE: {result['filename']}\n")
            f.write(f"{'='*80}\n\n")
            
            for sheet in result["sheets"]:
                f.write(f"--- Sheet: {sheet['name']} ---\n")
                f.write(f"Rows: {sheet['rows']}, Columns: {sheet['cols']}\n")
                
                if sheet["merged_cells"]:
                    f.write(f"Merged cells: {', '.join(sheet['merged_cells'][:10])}\n")
                
                f.write(f"\nHeaders (first 3 rows):\n")
                for i, row in enumerate(sheet["headers"], 1):
                    f.write(f"  Row {i}: {' | '.join(row)}\n")
                
                f.write(f"\nSample data (next 5 rows):\n")
                for i, row in enumerate(sheet["sample_rows"], 4):
                    f.write(f"  Row {i}: {' | '.join(row)}\n")
                
                if sheet["last_rows"]:
                    f.write(f"\nLast rows:\n")
                    for i, row in enumerate(sheet["last_rows"]):
                        f.write(f"  Row {sheet['rows']-len(sheet['last_rows'])+i+1}: {' | '.join(row)}\n")
                
                f.write(f"\n")
            
            f.write(f"\n")
    
    print(f"Report written to: {OUTPUT_FILE}")

if __name__ == "__main__":
    results = []
    for fname in sorted(os.listdir(RAW_DIR)):
        if fname.endswith(".xlsx"):
            fpath = os.path.join(RAW_DIR, fname)
            print(f"Analyzing: {fname}")
            results.append(analyze_workbook(fpath))
    
    write_report(results)
    print("Done!")
