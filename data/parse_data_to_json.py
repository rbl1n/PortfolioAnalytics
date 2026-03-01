import openpyxl
import json
import os
from datetime import datetime

# Paths
BASE_DIR = os.path.dirname(__file__)
RAW_DIR = os.path.join(BASE_DIR, "raw")
PROCESSED_DIR = os.path.join(BASE_DIR, "processed")
EXCEL_FILE = os.path.join(RAW_DIR, "理財彙總2025_珊.xlsx")
OUTPUT_JSON = os.path.join(PROCESSED_DIR, "data.json")

def parse_excel_to_json():
    print(f"Loading workbook: {EXCEL_FILE}")
    if not os.path.exists(EXCEL_FILE):
        print(f"Error: Could not find file at {EXCEL_FILE}")
        return

    # Ensure output directory exists
    os.makedirs(PROCESSED_DIR, exist_ok=True)

    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    monthly_sheets = [s for s in wb.sheetnames if s.startswith("各行理財存款清單-")]
    
    # Sort sheets by month
    monthly_sheets.sort(key=lambda x: x.split("-")[1])
    
    portfolio_data = {
        "months": [],
        "funds": {},     # Key: cert_id, Value: static fund metadata
        "history": {}    # Key: cert_id, Value: Dict of month -> performance metrics
    }

    for sheet_name in monthly_sheets:
        month = sheet_name.split("-")[1]
        portfolio_data["months"].append(month)
        ws = wb[sheet_name]
        
        # Determine header row (usually around row 4 or 5)
        header_row = None
        for r in range(1, 10):
            if ws.cell(row=r, column=1).value == "憑證號碼":
                header_row = r
                break
                
        if not header_row:
            print(f"Warning: Could not find header row in {sheet_name}. Skipping.")
            continue
            
        # Map headers to column indices
        headers = {}
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=header_row, column=col).value
            if val:
                headers[str(val).strip()] = col
                
        # Required columns mapping
        col_map = {
            "cert": headers.get("憑證號碼"),
            "name": headers.get("基金名稱"),
            "type": headers.get("型態"),
            "currency": headers.get("幣別"),
            "dividend_type": headers.get("配息"),
            "current_value": headers.get("現值"),
            "cumulative_dividend": headers.get("累計配息"),
            "profit_loss": headers.get("含息損益"),
            "profit_loss_ex_div": headers.get("不含息損益"),
            "return_rate": headers.get("含息報酬%"),
        }
        
        # Track current bank grouping
        current_bank = "其他"
        
        for row_idx in range(header_row + 1, ws.max_row + 1):
            col1_val = ws.cell(row=row_idx, column=1).value
            
            # Detect bank header (text in col1, empty in col2, not a cert number)
            if col1_val and not ws.cell(row=row_idx, column=2).value and not str(col1_val).startswith("00"):
                if "合計" not in str(col1_val) and "粗估" not in str(col1_val):
                    current_bank = str(col1_val).strip()
                continue
                
            # Process fund row
            if col1_val and str(col1_val).startswith("00"):
                try:
                    cert_id = str(col1_val).strip()
                    
                    # Store static metadata if not exists
                    if cert_id not in portfolio_data["funds"]:
                        portfolio_data["funds"][cert_id] = {
                            "cert": cert_id,
                            "bank": current_bank,
                            "name": str(ws.cell(row=row_idx, column=col_map["name"]).value).strip() if col_map["name"] else "Unknown",
                            "type": str(ws.cell(row=row_idx, column=col_map["type"]).value).strip() if col_map["type"] else "Unknown",
                            "currency": str(ws.cell(row=row_idx, column=col_map["currency"]).value).strip() if col_map["currency"] else "NTD",
                            "dividend_type": str(ws.cell(row=row_idx, column=col_map["dividend_type"]).value).strip() if col_map["dividend_type"] else "無"
                        }
                        portfolio_data["history"][cert_id] = {}
                        
                    # Store monthly historical data
                    def parse_float(val):
                        if val is None or val == "": return 0.0
                        try: return float(val)
                        except: return 0.0
                        
                    portfolio_data["history"][cert_id][month] = {
                        "current_value": parse_float(ws.cell(row=row_idx, column=col_map["current_value"]).value if col_map["current_value"] else 0),
                        "cumulative_dividend": parse_float(ws.cell(row=row_idx, column=col_map["cumulative_dividend"]).value if col_map["cumulative_dividend"] else 0),
                        "profit_loss": parse_float(ws.cell(row=row_idx, column=col_map["profit_loss"]).value if col_map["profit_loss"] else 0),
                        "profit_loss_ex_div": parse_float(ws.cell(row=row_idx, column=col_map["profit_loss_ex_div"]).value if col_map["profit_loss_ex_div"] else 0),
                        "return_rate": parse_float(ws.cell(row=row_idx, column=col_map["return_rate"]).value if col_map["return_rate"] else 0) * 100 # Convert to percentage
                    }
                except Exception as e:
                    print(f"Warning: Error processing row {row_idx} in {sheet_name}: {e}")

    # Calculate month-over-month trend (Up/Down/Flat)
    for cert_id, history in portfolio_data["history"].items():
        sorted_months = sorted(history.keys())
        for i in range(len(sorted_months)):
            month = sorted_months[i]
            if i == 0:
                history[month]["trend"] = "flat"
            else:
                prev_month = sorted_months[i-1]
                curr_val = history[month]["current_value"]
                prev_val = history[prev_month]["current_value"]
                
                if curr_val > prev_val:
                    history[month]["trend"] = "up"
                elif curr_val < prev_val:
                    history[month]["trend"] = "down"
                else:
                    history[month]["trend"] = "flat"

    # Write output JSON
    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(portfolio_data, f, ensure_ascii=False, indent=2)
        
    print(f"Successfully processed {len(portfolio_data['funds'])} funds across {len(portfolio_data['months'])} months.")
    print(f"Saved normalized data to: {OUTPUT_JSON}")

if __name__ == "__main__":
    parse_excel_to_json()
