"""Analyze the DU field across monthly sheets to determine its meaning."""
import openpyxl
import os

FILEPATH = os.path.join(os.path.dirname(__file__), "raw", "理財彙總2025_珊.xlsx")
OUTPUT = os.path.join(os.path.dirname(__file__), "du_analysis.txt")

def analyze_du():
    wb = openpyxl.load_workbook(FILEPATH, data_only=True)
    
    # Collect sheets that have DU column (the later months: 1231, 1130)
    target_sheets = [s for s in wb.sheetnames if s.startswith("各行理財存款清單")]
    
    results = []
    
    for sheet_name in target_sheets:
        ws = wb[sheet_name]
        
        # Find header row and DU column
        du_col = None
        header_row = None
        cert_col = None
        fund_col = None
        value_col = None  # 現值 column
        
        for row_idx in range(1, min(10, ws.max_row + 1)):
            for col_idx in range(1, ws.max_column + 1):
                cell_val = ws.cell(row=row_idx, column=col_idx).value
                if cell_val and "現值DU" in str(cell_val):
                    du_col = col_idx
                    header_row = row_idx
                if cell_val and str(cell_val) == "憑證號碼":
                    cert_col = col_idx
                    if header_row is None:
                        header_row = row_idx
                if cell_val and str(cell_val) == "基金名稱":
                    fund_col = col_idx
                if cell_val and str(cell_val) == "現值":
                    value_col = col_idx

        sheet_data = {
            "sheet": sheet_name,
            "du_col": du_col,
            "header_row": header_row,
            "entries": []
        }
        
        if header_row and cert_col:
            for row_idx in range(header_row + 1, ws.max_row + 1):
                cert = ws.cell(row=row_idx, column=cert_col).value
                if cert and str(cert).startswith("00"):
                    fund = ws.cell(row=row_idx, column=fund_col).value if fund_col else None
                    du_val = ws.cell(row=row_idx, column=du_col).value if du_col else None
                    val = ws.cell(row=row_idx, column=value_col).value if value_col else None
                    
                    sheet_data["entries"].append({
                        "cert": cert,
                        "fund": str(fund)[:30] if fund else "",
                        "du": du_val,
                        "value": val
                    })
        
        results.append(sheet_data)
    
    wb.close()
    
    # Write analysis
    with open(OUTPUT, "w", encoding="utf-8") as f:
        f.write("=== DU Field Analysis ===\n\n")
        
        # Check which sheets have DU column
        f.write("Sheets with DU column:\n")
        for r in results:
            has_du = "YES" if r["du_col"] else "NO"
            f.write(f"  {r['sheet']}: DU column = {has_du} (col {r['du_col']})\n")
        
        f.write("\n\n=== DU Values per Certificate ===\n\n")
        
        # Collect all unique certificates
        all_certs = set()
        for r in results:
            for e in r["entries"]:
                all_certs.add(e["cert"])
        
        # For each certificate, show DU values across months
        for cert in sorted(all_certs)[:20]:  # Limit to first 20
            f.write(f"\nCert: {cert}\n")
            for r in results:
                for e in r["entries"]:
                    if e["cert"] == cert:
                        f.write(f"  {r['sheet'][-4:]}: DU={e['du']}, Value={e['value']}, Fund={e['fund']}\n")
        
        # Summary of DU values
        f.write("\n\n=== DU Value Distribution ===\n")
        du_values = {}
        for r in results:
            for e in r["entries"]:
                dv = str(e["du"])
                if dv not in du_values:
                    du_values[dv] = 0
                du_values[dv] += 1
        
        for k, v in sorted(du_values.items()):
            f.write(f"  '{k}': {v} occurrences\n")
    
    print(f"Analysis written to {OUTPUT}")

if __name__ == "__main__":
    analyze_du()
